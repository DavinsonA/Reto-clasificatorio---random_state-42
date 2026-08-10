"""Parser de los Excel del corpus: datasets del AI Index (4 archivos).

Misma logica que csv_docs.py: una fila -> un bloque. Si el libro trae varias
hojas, cada bloque se prefija con el nombre de hoja para no perder esa senal
(las hojas del AI Index suelen ser cortes distintos del mismo dataset).

Tres columnas son identificadores tecnicos sin valor semantico (pmid, Author
ID, Conference ID de PubMed/Microsoft Academic Graph): se excluyen del texto
indexable, pero la fila sigue siendo un bloque con el resto de sus columnas.
"""

from __future__ import annotations

from typing import Any

from openpyxl import load_workbook

from .core import CatalogEntry, RawDoc, clean

# Whitelist explicita y pequena a proposito: una regla amplia tipo "id" in col
# excluiria columnas legitimas como "Conference Name" o futuras columnas con
# "id" en el nombre que si aporten texto.
EXCLUDED_COLUMNS = {"pmid", "author id", "conference id"}


def extract(entry: CatalogEntry) -> RawDoc:
    """Convierte un XLSX del corpus en `RawDoc`: una fila -> un bloque."""
    book = load_workbook(entry.path, read_only=True, data_only=True)
    try:
        hojas = [s for s in book.worksheets if s.sheet_state == "visible"] or book.worksheets
        multi_hoja = len(hojas) > 1
        blocks: list[str] = []
        columnas: dict[str, list[str]] = {}
        excluidas: dict[str, list[str]] = {}

        for sheet in hojas:
            rows = sheet.iter_rows(values_only=True)
            header = [clean(str(c)) if c is not None else "" for c in next(rows, ())]
            columnas[sheet.title] = header
            excluded = [col for col in header if col.casefold() in EXCLUDED_COLUMNS]
            if excluded:
                excluidas[sheet.title] = excluded

            for row in rows:
                block = _row_to_block(header, row)
                if block:
                    blocks.append(f"[{sheet.title}] {block}" if multi_hoja else block)
    finally:
        book.close()

    extra: dict[str, Any] = {
        "observatorio": entry.observatory,
        "tabular": True,
        "columnas": columnas,
        "num_filas": len(blocks),
    }
    if excluidas:
        extra["columnas_excluidas"] = excluidas

    if not blocks:
        blocks = [f"{entry.path.stem}. Observatorio: {entry.observatory}"]
        extra["contenido_minimo"] = True

    return RawDoc(
        doc_id=entry.doc_id,
        fuente=entry.fuente,
        formato=entry.formato,
        fenomeno=entry.fenomeno,
        title="",
        blocks=tuple(blocks),
        extra=extra,
    )


def _row_to_block(header: list[str], row: tuple[Any, ...]) -> str:
    """Formatea una fila como 'columna: valor | columna: valor', omitiendo vacios e IDs tecnicos."""
    pares = []
    for col, valor in zip(header, row):
        if not col or col.casefold() in EXCLUDED_COLUMNS:
            continue
        valor_limpio = clean(str(valor)) if valor not in (None, "") else ""
        if valor_limpio:
            pares.append(f"{col}: {valor_limpio}")
    return " | ".join(pares)
