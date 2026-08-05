"""Contratos de la extraccion y catalogo de ADL."""

from __future__ import annotations

import logging
import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

ROOT = Path(__file__).resolve().parents[2]
DATA_ROOT = ROOT / "data"
INDEX_PATH = DATA_ROOT / "Indice_Datos_Codefest.xlsx"

# Invisibles de categoria Cf: no son whitespace, `\s` no los toca.
_INVISIBLE = dict.fromkeys((0x00AD, 0x200B, 0x200C, 0x200D, 0x2060, 0xFEFF))
_CONTROL = re.compile(r"[\x00-\x08\x0b\x0e-\x1f\x7f-\x9f]")
_BLANKS = re.compile(r"\s+")


def clean(text: Any) -> str:
    """Normaliza a NFC, borra invisibles y controles y colapsa blancos."""
    if not isinstance(text, str) or not text:
        return ""
    text = unicodedata.normalize("NFC", text).translate(_INVISIBLE)
    return _BLANKS.sub(" ", _CONTROL.sub(" ", text)).strip()


@dataclass(frozen=True, slots=True)
class CatalogEntry:
    """Fila del indice de ADL resuelta contra el disco."""

    doc_id: str
    fuente: str  # ruta relativa, no nombre: 186 archivos comparten nombre (C1)
    formato: str  # extension real en minusculas (C2)
    fenomeno: int
    observatory: str
    path: Path


@dataclass(frozen=True, slots=True)
class RawDoc:
    """Texto de un documento sin fragmentar. Invariante: nunca cero bloques."""

    doc_id: str
    fuente: str
    formato: str
    fenomeno: int
    title: str
    blocks: tuple[str, ...]
    extra: dict[str, Any] = field(default_factory=dict)


def load_catalog(index_path: Path = INDEX_PATH, data_root: Path = DATA_ROOT) -> list[CatalogEntry]:
    """Lee el indice de ADL, que es la lista de entrada del pipeline y no `data/**`."""
    book = load_workbook(index_path, read_only=True, data_only=True)
    rows = book["Inventario de Archivos"].iter_rows(values_only=True)
    column = {name: position for position, name in enumerate(next(rows))}
    entries = []
    for row in rows:
        if not row[column["DOC_ID"]]:
            continue
        fuente = f"{row[column['Carpeta']]}/{row[column['Nombre estandarizado']]}"
        fuente = fuente.replace("\\", "/")
        entries.append(
            CatalogEntry(
                doc_id=str(row[column["DOC_ID"]]),
                fuente=fuente,
                formato=Path(fuente).suffix.lstrip(".").lower(),
                fenomeno=int(str(row[column["Fenómeno"]])[1]),
                observatory=str(row[column["Observatorio"]]),
                path=data_root.joinpath(*fuente.split("/")),
            )
        )
    book.close()

    found = [entry for entry in entries if entry.path.is_file()]
    if len(found) < len(entries):
        logger.warning("%d documentos del indice no estan en disco", len(entries) - len(found))
    logger.info("catalogo: %d documentos", len(found))
    return found
