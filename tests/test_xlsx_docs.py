"""Tests del parser de XLSX: exclusion de identificadores tecnicos."""

from __future__ import annotations

from openpyxl import Workbook

from src.extract import xlsx_docs


def _build_workbook(path) -> None:
    """Un libro con las 4 formas de columnas reales del corpus, en 4 hojas."""
    book = Workbook()

    lit_covid = book.active
    lit_covid.title = "lit_covid"
    lit_covid.append(["pmid", "title", "journal"])
    lit_covid.append([32634823, "Clinical observations in COVID-19", "Geriatr"])

    authors = book.create_sheet("authors")
    authors.append(["Author", "Author ID"])
    authors.append(["Andrew Y Ng", 2104401652])

    conferences = book.create_sheet("conferences")
    conferences.append(["Conference Name", "Conference ID"])
    conferences.append(["CVPR", 1158167855])

    fields = book.create_sheet("fields")
    fields.append(["Fields", "Status"])
    fields.append(["Artificial Intelligence", "Inprogress"])

    book.save(path)


def test_excluye_ids_tecnicos(tmp_path, make_entry):
    path = tmp_path / "AIINDEX_test.xlsx"
    _build_workbook(path)
    entry = make_entry(path, doc_id="F1-AIINDEX-042")

    doc = xlsx_docs.extract(entry)

    texto = " ".join(doc.blocks)
    assert "pmid:" not in texto
    assert "Author ID:" not in texto
    assert "Conference ID:" not in texto


def test_preserva_columnas_humanas(tmp_path, make_entry):
    path = tmp_path / "AIINDEX_test.xlsx"
    _build_workbook(path)
    entry = make_entry(path, doc_id="F1-AIINDEX-042")

    doc = xlsx_docs.extract(entry)

    texto = " ".join(doc.blocks)
    for campo in ("title:", "journal:", "Author:", "Conference Name:", "Fields:", "Status:"):
        assert campo in texto


def test_ejemplo_exacto_del_encargo(tmp_path, make_entry):
    path = tmp_path / "AIINDEX_test.xlsx"
    _build_workbook(path)
    entry = make_entry(path, doc_id="F1-AIINDEX-042")

    doc = xlsx_docs.extract(entry)

    assert "[authors] Author: Andrew Y Ng" in doc.blocks
    assert "[conferences] Conference Name: CVPR" in doc.blocks


def test_multi_hoja_no_se_rompe_y_sin_bloques_vacios(tmp_path, make_entry):
    path = tmp_path / "AIINDEX_test.xlsx"
    _build_workbook(path)
    entry = make_entry(path, doc_id="F1-AIINDEX-042")

    doc = xlsx_docs.extract(entry)

    assert len(doc.blocks) == 4  # una fila de datos por cada una de las 4 hojas
    assert all(block.strip() for block in doc.blocks)


def test_columnas_excluidas_quedan_en_extra(tmp_path, make_entry):
    path = tmp_path / "AIINDEX_test.xlsx"
    _build_workbook(path)
    entry = make_entry(path, doc_id="F1-AIINDEX-042")

    doc = xlsx_docs.extract(entry)

    assert doc.extra["columnas_excluidas"] == {
        "lit_covid": ["pmid"],
        "authors": ["Author ID"],
        "conferences": ["Conference ID"],
    }
